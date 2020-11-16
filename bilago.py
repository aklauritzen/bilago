import os
import glob
import io
import re
import time

import PySimpleGUI as sg
from PyPDF2 import PdfFileReader, PdfFileWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import red
from reportlab.lib.colors import white
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fills import PatternFill
from docx2pdf import convert

# Regular expression: content between { }
annex_number_pattern = '{(.*?)}'

# Regular expression: ABCD-12345-12345-20
journalnumber_pattern = r'[A-Za-z0-9]{4}-[A-Za-z0-9]{5}-[A-Za-z0-9]{5}-[A-Za-z0-9]{2}'

# Annex objects
annex_list = []

# Sourcefolder
source_folder = ''

# Destinationfolder
destination_folder = ''

# Annexcount
annex_count = 0

# Operationtitel
operation_titel = ''

# .docx files
docx_files_list = []

# Temporary converted .docx files
temporary_docx_pdf_files = []


class Annex(object):
    def __init__(self, journalnumber, base_filename, complete_filename, annex_number, num_pages):
        self.journalnumber = journalnumber
        self.base_filename = base_filename
        self.complete_filename = complete_filename
        self.annex_number = annex_number
        self.num_pages = num_pages


def inform_user():
    # Theme (Should always be fixed to default names: DarkBlack1 etc.)
    sg.LOOK_AND_FEEL_TABLE['DarkBlack1'] = {'BACKGROUND': '#f8f5f1',
                                            'TEXT': '#001e3c',
                                            'INPUT': '#fff',
                                            'TEXT_INPUT': '#000000',
                                            'SCROLL': '#99CC99',
                                            'BUTTON': ('#001e3c', '#d4e600'),
                                            'PROGRESS': ('#D1826B', '#CC8019'),
                                            'BORDER': 1, 'SLIDER_DEPTH': 0,
                                            'PROGRESS_DEPTH': 0, }
    sg.theme('DarkBlack1')

    layout = [[sg.Image(filename='static/images/header.png')],
              [sg.Text('Programbeskrivelse', font='"arial Bold" 13')],
              [sg.Text('Bilago opretter en kopi af dine PDF-filer og wordfiler hvorpå der påføres\n'
                       'bilagsnumre og sidetal. Bemærk at dine originale filer ikke ændres i processen.\n\n'
                       f'Resultatet findes efterfølgende i en ny mappe med titlen "Bilageret {datetime.now().date()}".\n'

                       'Bilagsnumre defineres ud fra indholdet mellem tuborgklammer i filnavnet.')],
              {sg.Text('Eksempler på filnavne med bilagsnumre', font='arial 12')},

              [sg.Text('▶', text_color='#001E3C', font='arial 15'), sg.Text('{4-01-1} - Rapport.pdf')],
              [sg.Text('▶', text_color='#001E3C', font='arial 15'),
               sg.Text('{A-1-2-1} - 13B Rapport Jens Jensen.docx')],
              [sg.Text('▶', text_color='#001E3C', font='arial 15'),
               sg.Text('{1-2} - Fotorapport Skovvej 31.pdf', font='arial 12')],
              [sg.Text('\nTuborgklammetegnet laves med tastekombination Alt Gr + 7 eller 0\n', font='arial 12')],
              [sg.Text('Vælg mappe med filer der skal behandles', font='"arial Bold" 12')],
              [sg.InputText(), sg.FolderBrowse('Vælg mappe', font='"arial Bold" 12 bold')],

              [sg.Text('\nIndstillinger (valgfri)', font='"arial Bold" 12')],
              [sg.Text('Operationstitel', size=(12, 1)), sg.InputText()],
              [sg.Checkbox('Bilagsoversigt', default=True, enable_events=True, key='build_annex_overview_event'),
               sg.Checkbox('Åben mappe med resultat', default=True, key='open_folder_in_explorer_event')],
              [sg.Checkbox('Samlet PDF-fil med alle bilag', default=False, key='create_combined_pdf_file_event')],
              [sg.Text('')],
              [sg.Button('Start bilagering', font='"arial Bold" 12 bold')],
              [sg.Text(
                  '\nUdviklet af Anders Koed Lauritzen                                                           Version 0.8',
                  font='arial 12 italic')]
              ]
    window = sg.Window('BILA{GO}', layout, size=(600, 720), font='Arial')

    # Event Loop
    while True:
        event, values = window.read()

        # Listener "windows closed""
        if event == sg.WIN_CLOSED:
            break

        # Listener "Start bilagering"
        elif event == 'Start bilagering':

            # Store value from inputfield
            global source_folder
            source_folder = values[1]

            # Continue if inputfield is not empty
            if source_folder != "":
                global operation_titel
                operation_titel = str(values[2])

                # Start tasktimer
                task_started = time.perf_counter()

                # Show popupanimation
                sg.popup_animated(image_source='static/images/gear_loader.png', no_titlebar=False, title='Afvent...')

                # sg.one_line_progress_meter('Behandler filer', 0, len(annex_list), key='file_progress_meter')
                # sg.one_line_progress_meter('file_progress_meter', annex_count, len(annex_list), 'key',
                #                            'Optional message')

                locate_docx_files()
                locate_pdf_files()

                global annex_list
                if annex_list:

                    apply_watermarks()

                    # Text regarding annex_overview should be empty, if annex_overview is unchecked
                    build_annex_overview_popup_text = ''

                    if values['build_annex_overview_event']:
                        build_annex_overview()
                        build_annex_overview_popup_text = '▶ Bilagsoversigt genereret\n'

                    # Delete temporary converted .docx files
                    delete_temporary_converted_docx_files()

                    build_combined_pdf_created_text = ''
                    if values['create_combined_pdf_file_event']:
                        create_combined_pdf_file()
                        build_combined_pdf_created_text = '▶ Samlet PDF-fil genereret\n'

                    # Stop tasttimer
                    task_ended = time.perf_counter()

                    # Close popupanimation
                    sg.PopupAnimated(image_source=None)

                    # Calculate total task time
                    task_total_time = task_ended - task_started

                    global annex_count
                    current_annex_count = annex_count

                    global destination_folder
                    if values['open_folder_in_explorer_event']:
                        os.startfile(destination_folder)

                    sg.popup(f'Bilagering gennemført\n\n'
                             f'▶ {current_annex_count} filer bilageret\n'
                             + build_annex_overview_popup_text + build_combined_pdf_created_text +
                             f'\nHandling gennemført på {task_total_time:0.2f} sekunder', title='Gennemført')

                    # Reset annexlist
                    annex_list = []

                    # Reset destinationfolder
                    destination_folder = ''

                    # reset annex_count
                    annex_count = 0

                    # Reset docx_files_list
                    global docx_files_list
                    docx_files_list = []

                    # Reset temporary_docx_pdf_file
                    global temporary_docx_pdf_files
                    temporary_docx_pdf_files = []

                else:
                    # Close loading popup animation
                    sg.PopupAnimated(image_source=None)

                    sg.popup(f'Du har valgt en mappe uden PDF-filer eller wordfiler.\n'
                             f'Filer som skal bilageres, skal indeholde tuborgklammer\n'
                             f'som beskrevet i programbeskrivelsen.', title='Fejlbesked')

            else:
                sg.popup(f'Du mangler at vælge en mappe', title='Fejlbesked')

    window.close()


def locate_docx_files():
    for docx_file in glob.iglob(source_folder + './**/*.docx', recursive=True):

        # If .docx filename contains { } and not 'Bilageret'
        try:
            annex_number = re.search(annex_number_pattern, docx_file).group(1)

            # Does not contain "Bilageret" in filename
            if "Bilageret" not in docx_file:
                # Convert .docx file to pdf file
                convert_docx_to_pdf(docx_file)

        except AttributeError:
            annex_number = None


def convert_docx_to_pdf(docx_file):
    convert(docx_file)

    # Delay to resolve a problem when converting small .docxfiles.
    time.sleep(1)

    temporary_docx_pdf_files.append(docx_file)


def locate_pdf_files():
    # Find PDF files
    for pdf_file in glob.iglob(source_folder + './**/*.pdf', recursive=True):

        # Look for {} in the filename
        try:
            annex_number = re.search(annex_number_pattern, pdf_file).group(1)

            # Trim whitespace. Ex: "1- 2-3" becomes "1-2-3"
            annex_number = re.sub(r"\s+", "", annex_number)

            # Filename does not contain "bilageret"
            if "Bilageret" not in pdf_file:
                # Contains a journalnummer
                journalnumber = re.search(journalnumber_pattern, pdf_file)

                annex_list.append(Annex(journalnumber, os.path.basename(pdf_file), pdf_file, annex_number, None))

        except AttributeError:
            annex_number = None


def apply_watermarks():
    now = datetime.now()

    # Create folder for outputfiles.
    global destination_folder
    destination_folder = source_folder + f" - Bilageret {now.date()} - {now.strftime('%H%M%S')}"
    os.mkdir(destination_folder)

    for annex in annex_list:
        num_pages = apply_watermark_to_pdf_file(annex.base_filename,
                                                annex.complete_filename,
                                                annex.annex_number)

        # Receives pagecount from apply_watermark_to_pdf_file()
        annex.num_pages = num_pages

        global annex_count
        annex_count += 1


def apply_watermark_to_pdf_file(base_filename, complete_filename, annex_number):
    # Original fil (rb = open file for reading)
    original_pdf = open(complete_filename, 'rb')

    # Error when using Strict: "Xref table not zero-indexed. ID numbers for objects will be corrected"
    pdf_reader = PdfFileReader(original_pdf, strict=False)

    packet = io.BytesIO()
    can = canvas.Canvas(packet)

    for i in range(pdf_reader.numPages):
        page_num = can.getPageNumber()
        can.setFont("Helvetica", 20)
        text = f'{annex_number}-{format(page_num, "03d")}'

        # Count number of characters and multiply to calculate the width. ( x * 10 )
        characters_in_annex_number = len(text)

        page_center = (int(pdf_reader.pages[i].mediaBox[2]) / 2) - (characters_in_annex_number * 5)
        page_top = int(pdf_reader.pages[i].mediaBox[3]) - 30

        can.setFillColor(white)
        can.rect(page_center, page_top - 8, characters_in_annex_number * 10, 30, fill=1, stroke=0)

        can.setFillColor(red)
        can.drawString(page_center, page_top, text)
        can.showPage()

    can.save()
    packet.seek(0)
    pdf_watermark_reader = PdfFileReader(packet)

    pdf_writer = PdfFileWriter()

    for pageNum in range(0, pdf_reader.numPages):
        original_page = pdf_reader.getPage(pageNum)
        original_page.mergePage(pdf_watermark_reader.getPage(pageNum))
        pdf_writer.addPage(original_page)

    # wb = create file for writing
    # Add "Bilageret" at the end of the filename
    index = base_filename.find('.pdf')
    base_filename_with_bilageret = base_filename[:index] + f' - Bilageret {datetime.now().date()}' + base_filename[
                                                                                                     index:]

    result_pdf = open(os.path.join(destination_folder, base_filename_with_bilageret), 'wb')
    pdf_writer.write(result_pdf)

    # Close PDF file
    original_pdf.close()

    # Return number of pages
    return pdf_reader.numPages


def build_annex_overview():
    workbook = Workbook()
    workbook.create_sheet('Bilagsoversigt', 0)
    sheet = workbook.active

    sheet['A1'] = 'Bilag'
    sheet['B1'] = 'Rapport'
    sheet['C1'] = 'Sideantal'

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 9

    # Header layout
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

    # Insert data from annex_list
    for row in range(0, len(annex_list)):
        sheet.cell(column=1, row=row + 2, value=annex_list[row].annex_number)
        sheet.cell(column=2, row=row + 2, value=annex_list[row].base_filename.
                   replace('.pdf', '')
                   .replace('{' + annex_list[row].annex_number + '} - ', ''))
        sheet.cell(column=3, row=row + 2, value=annex_list[row].num_pages)

    # Conditional statement - Show doublets in "bilagsnumre"
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    duplicate_rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)
    sheet.conditional_formatting.add(f'A1:A{len(annex_list) + 1}', duplicate_rule)

    # Printsettings
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = True
    sheet.page_setup.fitToHeight = False
    sheet.oddHeader.center.text = operation_titel
    sheet.oddFooter.center.text = "Side &[Page] af &N"
    sheet.print_area = 'A:C'

    # Sorter bilagskolonnen
    # sheet.auto_filter.add_sort_condition(f'A1:A{len(annex_list) + 1}')
    # sheet.auto_filter.add_sort_condition('A:A')

    global destination_folder
    annex_overview_filename = destination_folder + f'/Bilagsoversigt {datetime.now().date()}.xlsx'
    workbook.save(filename=annex_overview_filename)


def create_combined_pdf_file():
    # Find all PDF files in the output folder
    for pdf_file in glob.iglob(destination_folder + '**/*.pdf', recursive=True):
        # Combine PDF files
        print(pdf_file.title())


def delete_temporary_converted_docx_files():
    for file in temporary_docx_pdf_files:
        pdf_filename = file.replace('.docx', '.pdf')
        os.remove(pdf_filename)


if __name__ == "__main__":
    inform_user()
